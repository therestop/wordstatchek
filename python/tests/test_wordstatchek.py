from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from wordstatchek.api import RetryPolicy, WordstatApiError, WordstatClient
from wordstatchek.checkpoint import CheckpointStore
from wordstatchek.export import export_results
from wordstatchek.models import CheckResult
from wordstatchek.phrases import load_phrases


class FakeResponse:
    def __init__(self, status: int, payload: dict[str, object], *, retry_after: str | None = None) -> None:
        self.status_code = status
        self._payload = payload
        self.headers = {"Retry-After": retry_after} if retry_after else {}
        self.reason = "test"
        self.text = json.dumps(payload)

    def json(self) -> dict[str, object]:
        return self._payload


class FakeSession:
    def __init__(self, responses: list[FakeResponse]) -> None:
        self.responses = iter(responses)
        self.calls: list[dict[str, object]] = []

    def post(self, url: str, **kwargs: object) -> FakeResponse:
        self.calls.append({"url": url, **kwargs})
        return next(self.responses)


def test_phrase_file_removes_blanks_and_duplicates(tmp_path: Path) -> None:
    source = tmp_path / "input.txt"
    source.write_text("\ufeffсварка\n\n  сварка  \nметалл\n", encoding="utf-8")
    assert load_phrases(source) == ["сварка", "металл"]


def test_checkpoint_roundtrip(tmp_path: Path) -> None:
    store = CheckpointStore(tmp_path / "state.json")
    source = {"сварка": CheckResult("сварка", 42, "nonzero")}
    store.save(source)
    loaded = store.load()
    assert loaded["сварка"].total_count == 42


def test_api_retries_limit_and_uses_api_key_header() -> None:
    session = FakeSession([
        FakeResponse(429, {"message": "limit"}, retry_after="0"),
        FakeResponse(200, {"totalCount": "17"}),
    ])
    delays: list[float] = []
    client = WordstatClient(
        "secret",
        "folder",
        retry=RetryPolicy(attempts=2, base_delay=0, max_delay=0),
        session=session,  # type: ignore[arg-type]
        sleeper=delays.append,
    )
    assert client.total_count("фраза") == 17
    assert session.calls[0]["headers"]["Authorization"] == "Api-Key secret"  # type: ignore[index]
    assert session.calls[0]["json"]["folderId"] == "folder"  # type: ignore[index]
    assert delays == [0]


def test_api_marks_auth_errors_as_fatal() -> None:
    client = WordstatClient(
        "bad",
        "folder",
        session=FakeSession([FakeResponse(403, {"message": "denied"})]),  # type: ignore[arg-type]
    )
    with pytest.raises(WordstatApiError) as caught:
        client.total_count("фраза")
    assert caught.value.fatal is True


def test_export_creates_machine_and_human_readable_files(tmp_path: Path) -> None:
    values = {
        "one": CheckResult("one", 10, "nonzero"),
        "two": CheckResult("two", 0, "zero"),
        "three": CheckResult("three", None, "error", "boom"),
    }
    export_results(values, tmp_path)
    expected = {
        "wordstat_all.csv",
        "wordstat_nonzero.csv",
        "wordstat_nonzero.txt",
        "wordstat_zero.txt",
        "wordstat_errors.txt",
        "wordstat_results.xlsx",
    }
    assert expected <= {path.name for path in tmp_path.iterdir()}
    workbook = load_workbook(tmp_path / "wordstat_results.xlsx", read_only=True)
    assert workbook.sheetnames == ["Все", "Ненулевые", "Нулевые", "Ошибки"]
